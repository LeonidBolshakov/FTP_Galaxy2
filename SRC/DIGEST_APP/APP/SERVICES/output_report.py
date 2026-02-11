from pathlib import Path
import os

from typing import Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font

from SRC.DIGEST_APP.APP.dto import RuntimeContext, DescriptionOfNewTask
from SRC.DIGEST_APP.CONFIG.config import FontConfig, AlignmentConfig, ColumnConfig


class OutputReport:
    def run(
            self, ctx: RuntimeContext, descriptions: list[DescriptionOfNewTask]
    ) -> None:

        wb, ws = self.create_workbook_with_sheet(ctx=ctx, title="Дайджест обновлений")
        self.pack_head(ctx, ws)
        self.pack_info(ws, descriptions)
        self.tune_sheet(ctx, ws)
        self.close_worbook(ctx, wb, ws)

    def create_workbook_with_sheet(
            self, ctx: RuntimeContext, title: str
    ) -> tuple[Workbook, Worksheet]:
        work_book: Workbook = Workbook()
        sheet: Worksheet = work_book.create_sheet(title, index=0)

        return work_book, sheet

    def pack_head(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        columns: Sequence[ColumnConfig] = ctx.app.excel.columns
        ws.append([c.header for c in columns])

    def pack_info(
            self, ws: Worksheet, descriptions: list[DescriptionOfNewTask]
    ) -> None:
        for descr in descriptions:
            ws.append(
                [
                    descr.task,
                    ", ".join(descr.components),
                    descr.description,
                    descr.what_has_changed,
                    descr.how_it_changed,
                ]
            )

    def tune_sheet(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        self.tune_columns(ctx, ws)
        self.add_tune_head(ctx, ws)
        self.set_column_widths(ctx, ws)
        if ctx.app.excel.default.auto_filter:
            ws.auto_filter.ref = ws.dimensions

    def tune_columns(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        columns = ctx.app.excel.columns

        for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=len(columns),
        ):
            for cell, col_cfg in zip(row, columns, strict=False):
                self.tune_cell(cell, col_cfg.font, col_cfg.alignment)

    def add_tune_head(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        header_font = ctx.app.excel.header.font

        for cell in ws[1]:
            cell.font = Font(
                bold=header_font.bold,
            )

    def set_column_widths(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        for i, col_cfg in enumerate(ctx.app.excel.columns, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = col_cfg.width

    @staticmethod
    def tune_cell(cell, cell_font: FontConfig, cell_alignment: AlignmentConfig) -> None:

        cell.font = Font(
            name=cell_font.name,
            size=cell_font.size,
            bold=cell_font.bold,
        )

        cell.alignment = Alignment(
            horizontal=cell_alignment.horizontal,
            vertical=cell_alignment.vertical,
            wrapText=cell_alignment.wrap_text,
        )

    def close_worbook(
            self,
            ctx: RuntimeContext,
            wb: Workbook,
            ws: Worksheet,
    ) -> None:

        excel_path = ctx.app.excel.excel_path

        wb.save(excel_path)

        path = Path(excel_path).resolve()
        os.startfile(path)
