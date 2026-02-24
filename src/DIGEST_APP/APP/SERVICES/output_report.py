from pathlib import Path
import os
import sys
import subprocess

from typing import Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font

from DIGEST_APP.APP.dto import RuntimeContext, DescriptionOfNewTask
from DIGEST_APP.CONFIG.config import FontConfig, AlignmentConfig, ColumnConfig
from DIGEST_APP.APP.message import show_warning


class OutputReport:
    def run(
            self, ctx: RuntimeContext, descriptions: list[DescriptionOfNewTask]
    ) -> None:

        wb, ws = self._create_workbook_with_sheet(title="Дайджест обновлений")
        self._pack_head(ctx, ws)
        self._pack_info(ws, descriptions)
        self._tune_sheet(ctx, ws)
        self.close_worbook(ctx, wb, ws)

    def _create_workbook_with_sheet(self, title: str) -> tuple[Workbook, Worksheet]:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("В созданной книге EXCEL нет ни одной страницы. ")
        ws.title = title
        return wb, ws

    def _pack_head(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        columns: Sequence[ColumnConfig] = ctx.app.excel.columns
        ws.append([c.header for c in columns])

    def _pack_info(
            self, ws: Worksheet, descriptions: list[DescriptionOfNewTask]
    ) -> None:
        for descr in descriptions:
            ws.append(
                [
                    descr.task,
                    ", ".join(descr.components),
                    descr.description,
                    descr.what_has_changed,
                    descr.how_it_changed,
                ]
            )

    def _tune_sheet(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        self._tune_columns(ctx, ws)
        self._add_tune_head(ctx, ws)
        self._set_column_widths(ctx, ws)
        if ctx.app.excel.default.auto_filter:
            ws.auto_filter.ref = ws.dimensions

    def _tune_columns(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        columns = ctx.app.excel.columns

        for row in ws.iter_rows(
                min_row=1,
                max_row=ws.max_row,
                min_col=1,
                max_col=len(columns),
        ):
            for cell, col_cfg in zip(row, columns, strict=False):
                self._tune_cell(cell, col_cfg.font, col_cfg.alignment)

    def _add_tune_head(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        header_font = ctx.app.excel.header.font

        for cell in ws[1]:
            cell.font = Font(
                name=header_font.name,
                size=header_font.size,
                bold=header_font.bold,
            )

    def _set_column_widths(self, ctx: RuntimeContext, ws: Worksheet) -> None:
        for i, col_cfg in enumerate(ctx.app.excel.columns, start=1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = col_cfg.width

    @staticmethod
    def _tune_cell(
            cell, cell_font: FontConfig, cell_alignment: AlignmentConfig
    ) -> None:

        cell.font = Font(
            name=cell_font.name,
            size=cell_font.size,
            bold=cell_font.bold,
        )

        cell.alignment = Alignment(
            horizontal=cell_alignment.horizontal,
            vertical=cell_alignment.vertical,
            wrapText=cell_alignment.wrap_text,
        )

    def close_worbook(
            self,
            ctx: RuntimeContext,
            wb: Workbook,
            ws: Worksheet,
    ) -> None:

        excel_path = ctx.app.excel.excel_path

        try:
            wb.save(excel_path)
        except PermissionError as e:
            raise PermissionError(f"{excel_path.resolve()}") from e
        except OSError as e:
            raise OSError(f"Сохранение файла {excel_path.resolve()}") from e

        path = Path(excel_path).resolve()
        try:
            self.open_file(path)
        except (AttributeError, OSError) as e:
            show_warning(f"Файл сформирован и находится по пути {path}\n{e}")

    @staticmethod
    def open_file(path: str | Path) -> bool:
        """
        Пытается открыть файл системной программой по умолчанию.
        Возвращает True, если команда запуска была выполнена (без гарантии, что приложение реально открылось),
        и False, если открыть не удалось.
        Никогда не выбрасывает исключения наружу (подходит для продакшна и планировщика).
        """
        p = Path(path)

        # Если файла нет — просто "не смогли открыть"
        if not p.exists():
            return False

        try:
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path])
            elif sys.platform == "linux":
                subprocess.run(["xdg-open", path])
            else:
                return False
            return True
        except Exception:
            return False
